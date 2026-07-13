#!/usr/bin/env python
"""로컬 컨테이너에 붙어 create_excel을 실제로 돌려 end-to-end 검증한다.

- INFRA_MODE=local 강제, 웹훅/SSE는 스텁(실제 Discord 발송·에러 없음)
- 임시 Job을 만들어 export 실행 → 생성된 .xlsx를 다시 열어 행수/헤더 검증
- 끝나면 Job 행과 파일을 정리(단, Postgres 시퀀스 특성상 id 번호 한 칸은 빔)

사용:  cd backend && .venv/bin/python scripts/verify_export.py
전제:  로컬 Postgres(orders/jobs 테이블)가 떠 있어야 한다.
종료코드: 0=PASS, 1=FAIL
"""
import os
import sys
from time import perf_counter

# backend/ 를 import 경로에 추가 (scripts/ 에서 실행해도 절대 import 되도록)
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from sqlalchemy import func, select

# local 모드 강제 (GCS/Cloud Tasks 경로를 타지 않도록)
from core.config import Settings
Settings.INFRA_MODE = "local"

import excel.excel_service as es
from common.enums.job_status import JobStatus
from db.database import SessionLocal
from job.job import Job
from order.order import Order
import user.user  # noqa: F401 - Job.user_id FK(→users) 해석을 위해 메타데이터 등록


def main():
    # 외부 부작용 차단
    es.webhook_service.send_success_message = lambda *a, **k: None
    es.webhook_service.send_failure_message = lambda *a, **k: None
    es.publish_job_event = lambda *a, **k: None

    db = SessionLocal()
    total = db.scalar(select(func.count(Order.id))) or 0
    job = Job(status=JobStatus.PENDING, progress=0)
    db.add(job)
    db.commit()
    db.refresh(job)
    jid = job.id
    db.close()

    t = perf_counter()
    es.excel_service.create_excel(jid)
    elapsed = perf_counter() - t

    failures = []

    def check(name, cond):
        print(f"  [{'PASS' if cond else 'FAIL'}] {name}")
        if not cond:
            failures.append(name)

    db = SessionLocal()
    job = db.get(Job, jid)
    check(f"status == DONE ({job.status.value})", job.status == JobStatus.DONE)
    check(f"progress == 100 ({job.progress})", job.progress == 100)
    check(f"processed_rows == total ({job.processed_rows} == {total})", job.processed_rows == total)
    check("duration_seconds가 float", isinstance(job.duration_seconds, float))

    file_path = os.path.join(Settings().EXCEL_STORAGE_DIR, f"{jid}.xlsx")
    check("파일 생성됨", os.path.exists(file_path))
    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        rows = list(wb["Orders"].iter_rows(values_only=True))
        wb.close()
        check(f"행수 == header + {total} ({len(rows)})", len(rows) == total + 1)
        check("헤더 첫 칸 == 'ID'", rows and rows[0][0] == "ID")

    # 정리
    db.delete(job)
    db.commit()
    db.close()
    if os.path.exists(file_path):
        os.remove(file_path)

    print(f"\n총 {total}행, {elapsed:.2f}s, duration_seconds={job.duration_seconds}")
    if failures:
        print(f"RESULT: FAIL ({failures})")
        sys.exit(1)
    print("RESULT: PASS (정리 완료: job/파일 삭제)")


if __name__ == "__main__":
    main()
