"""create_excel end-to-end 흐름 (local 모드, SQLite + tmp 저장소).

성공 경로, 실패 시 FAILED 전이, 재시도 상한까지 실제 create_excel을 돌려
검증한다. 외부 부작용(웹훅/SSE/GCS)만 스텁한다.
"""
import pytest
from openpyxl import load_workbook

import excel.excel_service as es
from core.config import Settings
from common.enums.job_status import JobStatus
from job.job import Job
from order.order import Order


class _BoomStorage:
    """저장 단계에서 실패를 주입하는 스텁."""
    def save(self, *args, **kwargs):
        raise RuntimeError("storage boom")

    def upload(self, *args, **kwargs):
        raise RuntimeError("storage boom")


@pytest.fixture
def local_excel_env(monkeypatch, SessionFactory, tmp_path):
    # create_excel이 자체적으로 여는 세션을 테스트 엔진으로 돌린다.
    monkeypatch.setattr(es, "SessionLocal", SessionFactory)
    # local 모드 + tmp 저장 디렉토리
    monkeypatch.setattr(Settings, "INFRA_MODE", "local")
    monkeypatch.setattr(Settings, "EXCEL_STORAGE_DIR", str(tmp_path))
    # 외부 부작용 억제
    monkeypatch.setattr(es.webhook_service, "send_success_message", lambda *a, **k: None)
    monkeypatch.setattr(es.webhook_service, "send_failure_message", lambda *a, **k: None)
    monkeypatch.setattr(es, "publish_job_event", lambda *a, **k: None)
    return SessionFactory


def _seed(Session, order_count):
    s = Session()
    s.add_all([
        Order(user_name=f"u{i}", product_name="p", category="c", amount=i, status="PAID")
        for i in range(1, order_count + 1)
    ])
    job = Job(status=JobStatus.PENDING, progress=0)
    s.add(job)
    s.commit()
    jid = job.id
    s.close()
    return jid


def test_success_produces_valid_workbook(local_excel_env, tmp_path):
    Session = local_excel_env
    jid = _seed(Session, 12000)  # 5000 청크 → 3페이지

    es.excel_service.create_excel(jid)

    s = Session()
    job = s.get(Job, jid)
    assert job.status == JobStatus.DONE
    assert job.progress == 100
    assert job.processed_rows == 12000
    assert job.total_rows == 12000
    assert job.download_url
    s.close()

    # duration은 정수로 잘리지 않고 소수(2자리)로 보존된다
    assert isinstance(job.duration_seconds, float)

    fp = tmp_path / f"{jid}.xlsx"
    assert fp.exists()
    wb = load_workbook(fp, read_only=True)
    rows = list(wb["Orders"].iter_rows(values_only=True))
    wb.close()
    assert len(rows) == 12001  # header + 12000
    assert rows[0][0] == "ID"
    # keyset 페이징이 id 순서를 정확히 보존하는지 (1..12000, 누락/중복 없음)
    assert rows[1][0] == 1
    assert rows[-1][0] == 12000
    assert [r[0] for r in rows[1:]] == list(range(1, 12001))


def test_duration_seconds_keeps_two_decimals(local_excel_env, monkeypatch):
    """duration_seconds가 int로 잘리지 않고 소수 2자리로 반올림되는지 검증.

    create_excel 내부의 now() 호출을 통제된 시계로 대체한다. 현재 성공 경로에서
    now()는 3회 호출된다: (1) claim의 lease_cutoff, (2) claim의 started_at,
    (3) 완료 시 completed_at. started_at↔completed_at 차이를 3.146초로 두면
    round(.,2)=3.15 가 나와야 한다.
    """
    from datetime import datetime
    import excel.excel_service as es

    Session = local_excel_env
    jid = _seed(Session, 1)

    t_start = datetime(2026, 7, 13, 12, 0, 0)
    t_end = datetime(2026, 7, 13, 12, 0, 3, 146000)  # +3.146s

    class _Clock:
        def __init__(self, seq):
            self.seq = list(seq)
            self.i = 0

        def __call__(self):
            v = self.seq[min(self.i, len(self.seq) - 1)]
            self.i += 1
            return v

    monkeypatch.setattr(es, "now", _Clock([t_start, t_start, t_end]))

    es.excel_service.create_excel(jid)

    s = Session()
    job = s.get(Job, jid)
    assert job.duration_seconds == 3.15
    s.close()


def test_empty_orders_still_completes(local_excel_env, tmp_path):
    Session = local_excel_env
    jid = _seed(Session, 0)

    es.excel_service.create_excel(jid)

    s = Session()
    job = s.get(Job, jid)
    assert job.status == JobStatus.DONE
    assert job.progress == 100
    assert job.total_rows == 0
    s.close()


def test_failure_marks_failed_and_increments_attempt(local_excel_env, monkeypatch):
    Session = local_excel_env
    jid = _seed(Session, 1)
    monkeypatch.setattr(es, "get_storage_client", lambda: _BoomStorage())

    with pytest.raises(RuntimeError):
        es.excel_service.create_excel(jid)

    s = Session()
    job = s.get(Job, jid)
    assert job.status == JobStatus.FAILED
    assert job.attempt_count == 1
    assert job.error_message
    s.close()


def test_retry_cap_stops_after_max_attempts(local_excel_env, monkeypatch):
    Session = local_excel_env
    jid = _seed(Session, 1)
    monkeypatch.setattr(es, "get_storage_client", lambda: _BoomStorage())
    max_attempts = es.settings.MAX_ATTEMPTS

    # 매번 실패 → MAX_ATTEMPTS 회까지 재시도되며 attempt_count 증가
    for _ in range(max_attempts):
        with pytest.raises(RuntimeError):
            es.excel_service.create_excel(jid)

    s = Session()
    job = s.get(Job, jid)
    assert job.attempt_count == max_attempts
    assert job.status == JobStatus.FAILED
    s.close()

    # 상한 초과: claim이 안 되어 예외 없이 스킵되고 attempt_count도 그대로
    es.excel_service.create_excel(jid)  # 예외 발생하면 안 됨
    s = Session()
    job = s.get(Job, jid)
    assert job.attempt_count == max_attempts
    s.close()
