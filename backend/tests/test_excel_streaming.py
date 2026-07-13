"""write_only 스트리밍 워크북이 유효한 파일을 만드는지 라운드트립 검증."""
from openpyxl import Workbook, load_workbook


def test_write_only_round_trip(tmp_path):
    wb = Workbook(write_only=True)
    sheet = wb.create_sheet(title="Orders")
    header = ["ID", "User Name", "Product Name", "Category", "Amount", "Status", "Order date"]
    sheet.append(header)

    n = 5000
    for i in range(1, n + 1):
        sheet.append([i, f"user{i}", "prod", "cat", 100, "PAID", "2026-07-13 00:00:00"])

    fp = tmp_path / "out.xlsx"
    wb.save(fp)

    ro = load_workbook(fp, read_only=True)
    rows = list(ro["Orders"].iter_rows(values_only=True))
    ro.close()

    assert len(rows) == n + 1
    assert list(rows[0]) == header
    assert rows[-1][0] == n and rows[-1][1] == f"user{n}"
